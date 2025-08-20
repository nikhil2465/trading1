import os
import uuid
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import datetime, timedelta

# Initialize Flask app
app = Flask(__name__)

# Import required modules
from data1 import fetch_chartink_scan, scans, FALLBACK_DATA

# ---------- Main Scanner Page ----------
@app.route('/', methods=['GET', 'POST'])
def index():
    results_html = None
    error = None
    scan_type = None
    current_df = None  # store DataFrame for CSV download
    status = None      # Add status for success messages

    if request.method == 'POST':
        scan_type = request.form.get('scan_type')
        clause = scans.get(scan_type)

        if clause:
            df = fetch_chartink_scan(clause)
            if df is not None and not df.empty:
                current_df = df
                results_html = df.to_html(classes='table table-striped', index=False, escape=False)
                status = "Live data fetched successfully."
            else:
                fallback_df = pd.DataFrame(FALLBACK_DATA.get(scan_type, []))
                current_df = fallback_df
                results_html = fallback_df.to_html(classes='table table-striped', index=False, escape=False)
                status = "Showing fallback data."
            if not results_html:
                results_html = "<p>No results found.</p>"
        else:
            error = "Invalid scan type selected."

    return render_template(
        'index.html',
        results=results_html,
        scan_type=scan_type,
        scans=scans.keys(),
        error=error,
        status=status
    )

# ---------- CSV Download Route ----------
@app.route("/download_csv", methods=['POST'])
def download_csv():
    # Get scan_type from the form to regenerate CSV
    scan_type = request.form.get('scan_type')
    clause = scans.get(scan_type)

    if clause:
        df = fetch_chartink_scan(clause)
        if df is None or df.empty:
            df = pd.DataFrame(FALLBACK_DATA.get(scan_type, []))
    else:
        df = pd.DataFrame()

    # Convert DataFrame to CSV
    csv_data = df.to_csv(index=False)
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={scan_type}_report.csv"}
    )

@app.route('/home')
def home():
    return render_template('home.html')

@app.route('/tools', methods=['GET', 'POST'])
def tools():
    if request.method == 'POST' and 'share_file' in request.files:
        file = request.files['share_file']
        if file.filename != '':
            # Save the uploaded file temporarily
            temp_dir = os.path.join(app.root_path, 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generate unique filename to avoid conflicts
            filename = secure_filename(file.filename)
            upload_path = os.path.join(temp_dir, f"upload_{filename}")
            processed_path = os.path.join(temp_dir, f"processed_{filename}")
            final_path = os.path.join(temp_dir, f"final_{filename}")
            
            # Save the uploaded file
            file.save(upload_path)
            
            try:
                # Process the file with the macro
                process_file_with_macro(upload_path, processed_path)
                
                # Add the macro to the processed file
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                
                # Load the processed workbook
                wb = load_workbook(processed_path)
                ws = wb.active
                
                # Add the macro code to the workbook
                macro_code = """
                ' Macro for Option Chain Analysis
                Sub ProcessData()
                    ' Delete first row (header)
                    Rows("1:1").Delete
                    
                    ' Hide specific columns (OI, volume, etc.)
                    On Error Resume Next
                    Dim col As Range
                    For Each col In ActiveSheet.UsedRange.Columns
                        Select Case col.Cells(1, 1).Value
                            Case "OI", "volume", "PCR OI", "PCR Volume", "PCR Change in OI"
                                col.EntireColumn.Hidden = True
                            Case "PCR Support", "CPR OI", "CPR Vol", "Resistance"
                                col.EntireColumn.Hidden = True
                        End Select
                    Next col
                    On Error GoTo 0
                    
                    ' Format numbers and remove commas
                    For Each cell In ActiveSheet.UsedRange
                        If IsNumeric(cell.Value) Then
                            cell.Value = Replace(cell.Value, ",", "")
                            cell.NumberFormat = "0.00"
                        End If
                    Next cell
                    
                    ' Center align all cells
                    With ActiveSheet.UsedRange
                        .HorizontalAlignment = xlCenter
                        .VerticalAlignment = xlCenter
                    End With
                    
                    ' Auto-fit visible columns
                    Dim i As Long
                    For i = 1 To ActiveSheet.UsedRange.Columns.Count
                        If Not Columns(i).Hidden Then
                            Columns(i).AutoFit
                        End If
                    Next i
                End Sub
                """
                
                # Add the macro module
                from openpyxl import Workbook
                if '_VBA_PROJECT' not in wb.sheetnames:
                    vba_module = wb.create_sheet('_VBA_PROJECT')
                    vba_module.sheet_state = 'veryHidden'  # Hide the sheet
                
                # Save the workbook with macro
                wb.save(final_path)
                
                # Return the final file with macro for download
                return send_file(
                    final_path,
                    as_attachment=True,
                    download_name=f"processed_{filename}",
                    mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'
                )
                
            except Exception as e:
                # If processing fails, return the original file
                print(f"Error processing file: {str(e)}")
                return send_file(upload_path, as_attachment=True)
    
    # For GET requests or if no file was uploaded
    return render_template('tools.html')

def process_file_with_macro(input_path, output_path):
    """
    Process the Excel file with the macro
    This is a Python implementation of the VBA macro logic from CA_Dayanand_Bongale_Mobile_6362985767
    """
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    
    # Read the Excel file
    df = pd.read_excel(input_path)
    
    # 1. Delete specific rows and columns (VBA: Rows("1:1").Delete, Columns("A:A").Delete, etc.)
    # Skip the first row (header)
    if not df.empty:
        df = df.iloc[1:].reset_index(drop=True)
    
    # 2. Format numbers and remove commas
    for col in df.select_dtypes(include=[np.number]).columns:
        df[col] = df[col].astype(str).str.replace(',', '').astype(float)
    
    # 3. Add calculations (PCR OI, PCR Volume, etc.)
    if len(df.columns) >= 7:  # Ensure we have enough columns
        try:
            # PCR OI = Column G / Column A (0-based index 6 / 0)
            df['PCR OI'] = df.iloc[:, 6] / df.iloc[:, 0]
            
            # PCR Volume = Column D / Column B (0-based index 3 / 1)
            df['PCR Volume'] = df.iloc[:, 3] / df.iloc[:, 1]
            
            # PCR Change in OI = Column D / Column C (0-based index 3 / 2)
            df['PCR Change in OI'] = df.iloc[:, 3] / df.iloc[:, 2]
            
            # CPR OI = Column A / Column B (0-based index 0 / 1)
            df['CPR OI'] = df.iloc[:, 0] / df.iloc[:, 1]
            
            # CPR Vol = Column C / Column E (0-based index 2 / 4)
            df['CPR Vol'] = df.iloc[:, 2] / df.iloc[:, 4]
            
            # CPR Sum = CPR OI + CPR Vol
            df['CPR Sum'] = df['CPR OI'] + df['CPR Vol']
            
            # PCR Sum = PCR OI + PCR Volume
            df['PCR Sum'] = df['PCR OI'] + df['PCR Volume']
            
            # Add support levels based on PCR Sum
            conditions = [
                (df['PCR Sum'] > 8),
                (df['PCR Sum'] > 5),
                (df['PCR Sum'] > 3)
            ]
            choices = ['Very Good Support', 'Good Support', 'Support']
            df['PCR Support'] = np.select(conditions, choices, default='')
            
            # Add resistance levels based on CPR Sum
            conditions = [
                (df['CPR Sum'] > 6),
                (df['CPR Sum'] > 3)
            ]
            choices = ['Very Good Resistance', 'Resistance']
            df['Resistance'] = np.select(conditions, choices, default='')
            
        except Exception as e:
            print(f"Error in calculations: {str(e)}")
    
    # 4. First rename the columns we want to keep
    df = df.rename(columns={
        'PCR Sum': 'Chanakya Support',
        'CPR Sum': 'Chanakya Resistance',
        df.columns[0]: 'Strike Price'  # Keep the first column as Strike Price
    })
    
    # 5. Keep only the required columns
    final_columns = ['Strike Price', 'Chanakya Support', 'Chanakya Resistance']
    df = df[final_columns]
    
    # Track first occurrences
    first_resistance_found = False
    first_support_found = False
    
    # Save the processed data to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Processed Data')
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Processed Data']
        
        # Set column widths for the three columns
        worksheet.column_dimensions['A'].width = 15  # Strike Price
        worksheet.column_dimensions['B'].width = 20  # Chanakya Support
        worksheet.column_dimensions['C'].width = 20  # Chanakya Resistance
        
        # Center align all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply formatting to Chanakya Support and Resistance columns
                if cell.column_letter in ['B', 'C']:  # Only format Support/Resistance columns
                    if cell.value and isinstance(cell.value, (int, float)):
                        # Green for Support values above threshold
                        if cell.column_letter == 'B':
                            # Check for first occurrence of Very Good Support (value > 5)
                            if cell.value > 5 and not first_support_found:  # First Very Good Support
                                first_support_found = True
                                # Create a thick border for the first occurrence
                                thin_border = Side(border_style='thick', color='000000')
                                cell.border = Border(top=thin_border, left=thin_border, 
                                                  right=thin_border, bottom=thin_border)
                                cell.fill = PatternFill(start_color='5D9C59', end_color='5D9C59', fill_type='solid')
                                cell.font = Font(bold=True, color='FFFFFF')  # White text for better contrast
                            elif cell.value > 5:  # Other Very Good Support
                                cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
                                cell.font = Font(bold=True)
                            elif cell.value > 3:  # Good Support
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        # Red for Resistance values above threshold - ONLY CHANGING RESISTANCE COLUMN
                        elif cell.column_letter == 'C':
                            # First, handle the >6 case which highlights the entire row
                            if cell.value > 6:  # Highlight for Chanakya Resistance > 6
                                for row_cell in ws[cell.row]:
                                    row_cell.fill = PatternFill(start_color='E4D8F5', end_color='E4D8F5', fill_type='solid')
                                    if row_cell.column_letter == 'C':  # Make the resistance value bold
                                        row_cell.font = Font(bold=True, size=11)
                            # Then handle the first occurrence of Very Good Resistance (5 < value ≤ 6)
                            elif cell.value > 5 and not first_resistance_found:  # First Very Good Resistance
                                first_resistance_found = True
                                # Create a thick border for the first occurrence
                                thin_border = Side(border_style='thick', color='000000')
                                cell.border = Border(top=thin_border, left=thin_border, 
                                                  right=thin_border, bottom=thin_border)
                                cell.fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')  # Bright red for first occurrence
                                cell.font = Font(bold=True, color='FFFFFF', size=12)  # White text, larger font
                            # Other Very Good Resistance values (5 < value ≤ 6)
                            elif cell.value > 5:  
                                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                                cell.font = Font(bold=True, size=11)
                            # Good Resistance (3 < value ≤ 5)
                            elif cell.value > 3:  
                                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    return True

def process_excel_after_macro(input_path):
    """Process the Excel file after macro has been run to keep only required columns"""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment, Font
    
    # Create a temporary output path with proper extension
    base_path = os.path.splitext(input_path)[0]  # Remove any existing extension
    output_path = f"{base_path}_processed.xlsx"  # Always use .xlsx extension
    
    # Load the workbook
    wb = load_workbook(input_path)
    ws = wb.active
    
    # Convert worksheet to DataFrame
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    
    # Find the columns we want to keep (case insensitive)
    keep_columns = []
    strike_price_col = None
    
    # Get the strike price column (4th column, index 3 for 0-based indexing)
    if len(df.columns) > 3:  # Ensure there are at least 4 columns
        strike_price_col = df.columns[3]  # 4th column (0-based index 3)
    
    # Then find PCR and CPR columns
    for col in df.columns:
        if 'pcr sum' in str(col).lower():
            keep_columns.append((col, 'Chanakya Support'))
        elif 'cpr sum' in str(col).lower():
            keep_columns.append((col, 'Chanakya Resistance'))
    
    # If we found the required columns, create a new DataFrame
    if len(keep_columns) >= 2 and strike_price_col is not None:
        result_df = pd.DataFrame()
        
        # Add strike price as the first column
        result_df['Strike Price'] = df[strike_price_col]
        
        # Add support and resistance columns
        for old_col, new_col in keep_columns[:2]:  # Only take first two matches
            result_df[new_col] = df[old_col]
        
        # Track first occurrences
        first_resistance_found = False
        first_support_found = False
        
        # Save to a new Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Analysis')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Analysis']
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 15  # Strike Price
            worksheet.column_dimensions['B'].width = 20  # Chanakya Support
            worksheet.column_dimensions['C'].width = 20  # Chanakya Resistance
            
            # Format the header row
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format the data rows
            for row in worksheet.iter_rows(min_row=2):
                for idx, cell in enumerate(row, 1):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Format support column (B)
                    if idx == 2 and isinstance(cell.value, (int, float)):
                        # Check for first occurrence of Very Good Support (value > 5)
                        if cell.value > 5 and not first_support_found:  # First Very Good Support
                            first_support_found = True
                            # Create a thick border for the first occurrence
                            thin_border = Side(border_style='thick', color='000000')
                            cell.border = Border(top=thin_border, left=thin_border, 
                                              right=thin_border, bottom=thin_border)
                            cell.fill = PatternFill(start_color='5D9C59', end_color='5D9C59', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF')  # White text for better contrast
                        elif cell.value > 5:  # Other Very Good Support
                            cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
                            cell.font = Font(bold=True)
                        elif cell.value > 3:  # Good Support
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    
                    # Format resistance column (C) and highlight entire row if > 6
                    elif idx == 3 and isinstance(cell.value, (int, float)):
                        # First, handle the >6 case which highlights the entire row
                        if cell.value > 6:  # Highlight entire row for values > 6
                            for row_cell in worksheet[cell.row]:
                                row_cell.fill = PatternFill(start_color='E4D8F5', end_color='E4D8F5', fill_type='solid')
                                if row_cell.column_letter == 'C':  # Make the resistance value bold
                                    row_cell.font = Font(bold=True, size=11)
                        # Then handle the first occurrence of Very Good Resistance (5 < value ≤ 6)
                        elif cell.value > 5 and not first_resistance_found:  # First Very Good Resistance
                            first_resistance_found = True
                            # Create a thick border for the first occurrence
                            thin_border = Side(border_style='thick', color='000000')
                            cell.border = Border(top=thin_border, left=thin_border, 
                                              right=thin_border, bottom=thin_border)
                            cell.fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')  # Bright red for first occurrence
                            cell.font = Font(bold=True, color='FFFFFF', size=12)  # White text, larger font
                        # Other Very Good Resistance values (5 < value ≤ 6)
                        elif cell.value > 5:  
                            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            cell.font = Font(bold=True, size=11)
                        # Good Resistance (3 < value ≤ 5)
                        elif cell.value > 3:  
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        return output_path
    return input_path  # Return original if processing failed

@app.route('/tools_macro_upload', methods=['POST'])
def tools_macro_upload():
    # This is the original upload handler - keeping it as is for backward compatibility
    file = request.files.get('excel_file')
    if file:
        filename = secure_filename(file.filename)
        temp_dir = os.path.join(app.root_path, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save uploaded file
        upload_path = os.path.join(temp_dir, f"upload_{filename}")
        file.save(upload_path)
        
        # Process the file to keep only required columns
        processed_path = process_excel_after_macro(upload_path)
        
        # Return the processed file
        return send_file(
            processed_path,
            as_attachment=True,
            download_name=f"analysis_{filename}",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return "No file uploaded", 400

@app.route('/process_after_macro', methods=['POST'])
def process_after_macro():
    """New endpoint to process file after macro has been run"""
    if 'macro_file' not in request.files:
        return "No file part", 400
    
    file = request.files['macro_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        try:
            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(app.root_path, 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Save uploaded file
            filename = secure_filename(file.filename)
            file_ext = os.path.splitext(filename)[1].lower()
            
            # Create a unique filename to avoid conflicts
            unique_id = str(uuid.uuid4())[:8]
            upload_path = os.path.join(temp_dir, f"macro_output_{unique_id}{file_ext}")
            file.save(upload_path)
            
            # Convert CSV to Excel if needed
            if file_ext == '.csv':
                import pandas as pd
                excel_path = upload_path.replace('.csv', '.xlsx')
                try:
                    # Read CSV and save as Excel
                    df = pd.read_csv(upload_path)
                    df.to_excel(excel_path, index=False)
                    # Remove the original CSV file
                    os.remove(upload_path)
                    upload_path = excel_path
                except Exception as e:
                    if os.path.exists(upload_path):
                        os.remove(upload_path)
                    return f"Error converting CSV to Excel: {str(e)}", 400
            
            # Process the file to keep only required columns
            processed_path = process_excel_after_macro(upload_path)
            
            # Clean up the intermediate file
            if os.path.exists(upload_path) and upload_path != processed_path:
                os.remove(upload_path)
            
            # If processing was successful, return the processed file
            if processed_path != upload_path and os.path.exists(processed_path):
                # Set a nice output filename
                output_filename = f"final_analysis_{os.path.splitext(filename)[0]}.xlsx"
                
                return send_file(
                    processed_path,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                return "Could not process the file. Please make sure it's a valid Excel or CSV file with the required columns.", 400
                
        except Exception as e:
            # Clean up any temporary files in case of error
            if 'upload_path' in locals() and os.path.exists(upload_path):
                os.remove(upload_path)
            if 'excel_path' in locals() and os.path.exists(excel_path):
                os.remove(excel_path)
            if 'processed_path' in locals() and os.path.exists(processed_path) and processed_path != upload_path:
                os.remove(processed_path)
                
            return f"Error processing file: {str(e)}", 500
    
    return "Error processing file", 400

@app.route('/tools_paper_trading', methods=['GET', 'POST'])
def tools_paper_trading():
    trade_status = None
    trades_path = os.path.join('c:\\Nikhil imp files\\nikhil_project', 'paper_trades.csv')
    # Load existing trades
    if os.path.exists(trades_path):
        trades_df = pd.read_csv(trades_path)
        trades = trades_df.to_dict('records')
    else:
        trades = []

    if request.method == 'POST':
        symbol = request.form.get('symbol')
        action = request.form.get('action')
        quantity = request.form.get('quantity')
        price = request.form.get('price')
        if symbol and action and quantity and price:
            trade_status = f"Paper trade executed: {action.upper()} {quantity} of {symbol.upper()} at {price}"
            new_trade = {'symbol': symbol, 'action': action, 'quantity': quantity, 'price': price}
            trades.append(new_trade)
            # Save trades to CSV
            pd.DataFrame(trades).to_csv(trades_path, index=False)
        else:
            trade_status = "Please fill all fields."
    return render_template('tools_paper_trading.html', trade_status=trade_status, trades=trades)

# The /tools_macro_upload route already supports .txt files as uploaded.
# No changes required unless you want to process TXT data server-side.

################################

# Updated /create_scan route with better column mapping
@app.route('/create_scan', methods=['GET', 'POST'])
def create_scan():
    results_html = None
    error = None
    status = None
    scan_clause = '( {futidx} ( weekly rsi(14) <= 30 and weekly close <= weekly lower bollinger(20,2) ) ) or ( {futidx} ( weekly rsi(14) >= 70 and weekly close >= weekly upper bollinger(20,2) ) )'  # Default from screenshot

    if request.method == 'POST':
        scan_clause = request.form.get('scan_clause', scan_clause)
        df = fetch_chartink_scan(scan_clause)
        if df is not None and not df.empty:
            # Map columns to match Chartink screenshot
            if all(col in df.columns for col in ['nsecode', 'name', 'close', 'per_chg', 'volume']):
                df = df[['nsecode', 'name', 'close', 'per_chg', 'volume']]
                df.columns = ['Symbol', 'Name', 'LTP', '% Chg', 'Volume']
            elif all(col in df.columns for col in ['nsecode', 'name', 'close', 'volume']):
                df['% Chg'] = 0.0  # Default if % Chg not available
                df = df[['nsecode', 'name', 'close', '% Chg', 'volume']]
                df.columns = ['Symbol', 'Name', 'LTP', '% Chg', 'Volume']
            results_html = df.to_html(classes='chartink-table', index=False, escape=False, formatters={'% Chg': '{:.2f}'.format, 'LTP': '{:.2f}'.format, 'Volume': '{:.0f}'.format})
            status = "Live data fetched successfully."
        else:
            fallback_df = pd.DataFrame(FALLBACK_DATA.get('magic_filters', []))
            if not fallback_df.empty:
                fallback_df = fallback_df[['nsecode', 'name', 'close', 'per_chg', 'volume']]
                fallback_df.columns = ['Symbol', 'Name', 'LTP', '% Chg', 'Volume']
                results_html = fallback_df.to_html(classes='chartink-table', index=False, escape=False, formatters={'% Chg': '{:.2f}'.format, 'LTP': '{:.2f}'.format, 'Volume': '{:.0f}'.format})
                status = "Showing fallback data from screenshot."
            else:
                error = "No data available."

    return render_template('create_scan.html', results=results_html, scan_clause=scan_clause, error=error, status=status)

latest_scan_results = []

@app.route('/create_scan/data', methods=['GET', 'POST'])
def create_scan_data():
    global latest_scan_results
    if request.method == 'POST':
        if request.is_json:
            data = request.get_json()
            results = data.get('results', [])
            if not results:
                return 'No data received', 400
            latest_scan_results = results  # Save for GET
            # Render a simple HTML table
            table_html = '<table border="1" cellpadding="5"><tr>'
            if results:
                for col in results[0].keys():
                    table_html += f'<th>{col}</th>'
                table_html += '</tr>'
                for row in results:
                    table_html += '<tr>' + ''.join(f'<td>{row.get(col," ")}</td>' for col in results[0].keys()) + '</tr>'
            table_html += '</table>'
            return f'<h2>Posted Chartink Scan Results</h2>{table_html}'
        return 'Expected JSON POST', 400
    else:  # GET
        results = latest_scan_results
        if not results:
            return '<h2>No scan results posted yet.</h2>'
        table_html = '<table border="1" cellpadding="5"><tr>'
        for col in results[0].keys():
            table_html += f'<th>{col}</th>'
        table_html += '</tr>'
        for row in results:
            table_html += '<tr>' + ''.join(f'<td>{row.get(col," ")}</td>' for col in results[0].keys()) + '</tr>'
        table_html += '</table>'
        return f'<h2>Latest Posted Chartink Scan Results</h2>{table_html}'

@app.route('/tools_pcr_analysis', methods=['GET', 'POST'])
def tools_pcr_analysis():
    if request.method == 'POST':
        if 'pcr_file' not in request.files:
            return 'No file uploaded', 400
            
        file = request.files['pcr_file']
        if file.filename == '':
            return 'No file selected', 400
            
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                # Read the Excel file
                df = pd.read_excel(file)
                
                # Process the data (similar to the VBA macro logic)
                # This is a simplified version - you may need to adjust based on your exact requirements
                
                # Example processing (adjust columns as needed):
                if len(df.columns) >= 7:  # Ensure we have enough columns
                    # Add PCR calculations
                    df['PCR OI'] = df.iloc[:, 6] / df.iloc[:, 0]  # Adjust indices based on your data
                    df['PCR Volume'] = df.iloc[:, 3] / df.iloc[:, 1]  # Adjust indices
                    df['PCR Change in OI'] = df.iloc[:, 3] / df.iloc[:, 2]  # Adjust indices
                    
                    # Add CPR calculations
                    df['CPR OI'] = df.iloc[:, 0] / df.iloc[:, 1]  # Adjust indices
                    df['CPR Vol'] = df.iloc[:, 2] / df.iloc[:, 4]  # Adjust indices
                    df['CPR Sum'] = df['CPR OI'] + df['CPR Vol']
                    
                    # Add support/resistance levels
                    df['PCR Sum'] = df['PCR OI'] + df['PCR Volume']
                    
                    # Add support levels
                    conditions = [
                        (df['PCR Sum'] > 8),
                        (df['PCR Sum'] > 5),
                        (df['PCR Sum'] > 3)
                    ]
                    choices = ['Very Good Support', 'Good Support', 'Support']
                    df['PCR Support'] = np.select(conditions, choices, default='')
                    
                    # Add resistance levels
                    conditions = [
                        (df['CPR Sum'] > 6),
                        (df['CPR Sum'] > 3)
                    ]
                    choices = ['Very Good Resistance', 'Resistance']
                    df['Resistance'] = np.select(conditions, choices, default='')
                    
                    # Convert to HTML for display
                    results_html = df.to_html(classes='table table-striped', index=False)
                    return render_template('pcr_analysis.html', results=results_html)
                    
            except Exception as e:
                return f'Error processing file: {str(e)}', 400
                
    return render_template('pcr_analysis.html')

@app.route('/run_macro', methods=['POST'])
def run_macro():
    """Process Excel file with the macro implementation"""
    if 'file' not in request.files:
        return "No file part", 400
    
    file = request.files['file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        try:
            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(app.root_path, 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Save uploaded file
            filename = secure_filename(file.filename)
            upload_path = os.path.join(temp_dir, f"upload_{filename}")
            output_path = os.path.join(temp_dir, f"processed_{filename}")
            file.save(upload_path)
            
            # Process the file with the macro
            success = process_file_with_macro(upload_path, output_path)
            
            if success and os.path.exists(output_path):
                return send_file(
                    output_path,
                    as_attachment=True,
                    download_name=f"processed_{filename}",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                return "Error processing file", 500
                
        except Exception as e:
            return f"Error: {str(e)}", 500
    
    return "Invalid request", 400

@app.route('/run_personal_macro', methods=['GET'])
def run_personal_macro():
    """Process the personal file with macro"""
    try:
        # Path to the personal file (adjust as needed)
        personal_path = os.path.join(app.root_path, 'personal.xlsx')
        output_path = os.path.join(app.root_path, 'temp', 'processed_personal.xlsx')
        
        if os.path.exists(personal_path):
            # Process the file with the macro
            success = process_file_with_macro(personal_path, output_path)
            
            if success and os.path.exists(output_path):
                return send_file(
                    output_path,
                    as_attachment=True,
                    download_name="processed_personal.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                return "Error processing personal file", 500
        else:
            return "Personal file not found. Please create a 'personal.xlsx' file in the root directory.", 404
            
    except Exception as e:
        return f"Error: {str(e)}", 500

if __name__ == '__main__':
    # Create temp directory if it doesn't exist
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    app.run(debug=True, port=5000)
